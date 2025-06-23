import openpyxl
import pandas as pd
import tensorflow as tf
import random
import json
from sklearn.model_selection import train_test_split, StratifiedKFold
from sklearn.preprocessing import StandardScaler
from imblearn.over_sampling import SMOTE
from collections import Counter
from sklearn.metrics import accuracy_score, classification_report, confusion_matrix
import numpy as np
from tensorflow.keras import Sequential
from tensorflow.keras.layers import Conv1D, MaxPooling1D, Flatten, Dense, Dropout, Input
from tensorflow.keras.callbacks import EarlyStopping
from tensorflow.keras.regularizers import l2
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl.styles import Font, Alignment
from sklearn.decomposition import PCA
import time
import os
import joblib
from tensorflow.keras.models import load_model

# Cargar hiperparámetros
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BEST_PARAMS_PATH = os.path.join(BASE_DIR, 'best_params.json')
MODEL_PATH = os.path.join(BASE_DIR, 'modelo_cnn.h5')
SCALER_PATH = os.path.join(BASE_DIR, 'scaler.pkl')

with open(BEST_PARAMS_PATH, 'r') as f:
    best_params = json.load(f)

# Configuraciones globales
POOL_SIZE = best_params['pool_size']
BATCH_SIZE = best_params['batch_size']
EPOCHS = 30
PATIENCE = best_params['patience']

def establecer_semilla(seed=42):
    """Establece una semilla para reproducibilidad."""
    np.random.seed(seed)
    tf.random.set_seed(seed)
    random.seed(seed)

def comparar_con_pca(X_train, X_nuevos):
    """
    Compara datos de entrenamiento y nuevos datos usando PCA.

    Args:
        X_train (np.array): Datos de entrenamiento.
        X_nuevos (np.array): Nuevos datos.
    """
    X_train_2d = X_train.reshape(X_train.shape[0], X_train.shape[1])
    X_nuevos_2d = X_nuevos.reshape(X_nuevos.shape[0], X_nuevos.shape[1])
    X_total = np.vstack((X_train_2d, X_nuevos_2d))

    y_total = np.array([0] * X_train_2d.shape[0] + [1] * X_nuevos_2d.shape[0])
    pca = PCA(n_components=2)
    X_pca = pca.fit_transform(X_total)

    plt.figure(figsize=(10, 6))
    plt.scatter(X_pca[y_total == 0, 0], X_pca[y_total == 0, 1], label='Datos de Entrenamiento', alpha=0.6)
    plt.scatter(X_pca[y_total == 1, 0], X_pca[y_total == 1, 1], label='Nuevos Datos', alpha=0.6)
    plt.xlabel('Componente Principal 1')
    plt.ylabel('Componente Principal 2')
    plt.title('Comparación Entrenamiento vs Nuevos Datos (PCA)')
    plt.legend()
    plt.grid(True)
    plt.show()

def preparar_datos(df, test_size=0.2, random_state=42):
    """
    Preprocesa los datos: filtra, divide, balancea y prepara para la CNN.

    Args:
        df (pd.DataFrame): DataFrame con los datos.
        test_size (float): Proporción del conjunto de prueba.
        random_state (int): Semilla para reproducibilidad.

    Returns:
        tuple: X_train, X_test, y_train, y_test, scaler
    """

    # Filtrar filas donde ninguna banda (columnas 4 a 6) tenga un valor > 0.35
    bandas = df.iloc[:, 4:7]
    filtro = (bandas <= 0.35).all(axis=1)
    df_filtrado = df[filtro].copy()
    print(f"Filas originales: {len(df)}, Filas después de filtrar: {len(df_filtrado)}")

    # Preparar características (X) y etiquetas (y)
    X = df_filtrado.iloc[:, 4:].values
    y = df_filtrado['Especie'].apply(lambda x: 0 if x == 'PIC' else 1).values

    # Dividir en conjunto de entrenamiento y prueba
    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=test_size, random_state=random_state
    )

    # Contar clases en y_train antes de SMOTE
    train_counts = Counter(y_train)
    print(f'En y_train antes de SMOTE, 0: {train_counts[0]}, 1: {train_counts[1]}')

    # Aplicar SMOTE para balancear clases
    smote = SMOTE(random_state=random_state)
    X_train_2d = X_train.reshape(X_train.shape[0], -1)
    X_train_smote, y_train_smote = smote.fit_resample(X_train_2d, y_train)
    X_train = X_train_smote.reshape(X_train_smote.shape[0], X_train.shape[1], 1)
    y_train = y_train_smote

    # Contar clases después de SMOTE
    train_counts_smote = Counter(y_train)
    print(f'En y_train después de SMOTE, 0: {train_counts_smote[0]}, 1: {train_counts_smote[1]}')

    # Inicializar scaler
    scaler = StandardScaler()
    # Normalización comentada para pruebas
    # X_train = scaler.fit_transform(X_train.reshape(-1, X_train.shape[1])).reshape(X_train.shape)
    # X_test = scaler.transform(X_test.reshape(-1, X_test.shape[1])).reshape(X_test.shape[0], X_test.shape[1], 1)

    # Reshape X_test para CNN
    X_test = X_test.reshape(X_test.shape[0], X_test.shape[1], 1)

    print(f"Forma de X_train: {X_train.shape}, Forma de X_test: {X_test.shape}")
    return X_train, X_test, y_train, y_test, scaler

def crear_modelo(input_shape):
    """Crea y compila el modelo CNN."""
    model = Sequential()
    model.add(Input(shape=input_shape))
    model.add(Conv1D(filters=int(best_params['filters_1']), kernel_size=int(best_params['kernel_size']),
                     activation='elu', kernel_regularizer=l2(0.001)))
    model.add(MaxPooling1D(pool_size=POOL_SIZE))
    model.add(Conv1D(filters=int(best_params['filters_2']), kernel_size=int(best_params['kernel_size']),
                     activation='elu', kernel_regularizer=l2(0.001)))
    model.add(Conv1D(filters=int(best_params['filters_3']), kernel_size=int(best_params['kernel_size']),
                     activation='elu', kernel_regularizer=l2(0.001)))
    model.add(Conv1D(filters=int(best_params['filters_4']), kernel_size=int(best_params['kernel_size']),
                     activation='elu', kernel_regularizer=l2(0.001)))
    model.add(MaxPooling1D(pool_size=POOL_SIZE))
    model.add(Flatten())
    model.add(Dense(int(best_params['ff_dim']), activation='relu', kernel_regularizer=l2(0.001)))
    model.add(Dropout(0.5))
    model.add(Dense(1, activation='sigmoid'))

    model.compile(
        loss='binary_crossentropy',
        optimizer=tf.keras.optimizers.Adam(learning_rate=best_params['learning_rate']),
        metrics=['accuracy']
    )
    return model

def mostrar_resultados(y_true, y_pred, loss, accuracy, nombre_modelo, y_pred_prob, y_test):
    """
    Muestra métricas de evaluación del modelo.

    Args:
        y_true (np.array): Etiquetas reales.
        y_pred (np.array): Predicciones.
        loss (float): Pérdida del modelo.
        accuracy (float): Precisión del modelo.
        nombre_modelo (str): Nombre del modelo.
        y_pred_prob (np.array): Probabilidades predichas.
        y_test (np.array): Etiquetas de prueba.
    """
    print(f"Evaluación del modelo {nombre_modelo}")
    print(f"Pérdida: {loss:.4f}")
    print(f"Precisión: {accuracy:.4f}")
    print("\nInforme de clasificación:")
    print(classification_report(y_true, y_pred))

    cm = confusion_matrix(y_true, y_pred)
    plt.figure(figsize=(8, 6))
    sns.heatmap(cm, annot=True, fmt="d", cmap="YlGnBu", cbar=False,
                annot_kws={"size": 16, "weight": "bold"}, linewidths=1.5, linecolor="black")
    plt.title(f'Matriz de Confusión - {nombre_modelo}', fontsize=16, weight='bold')
    plt.xlabel('Predicciones', fontsize=14, weight='bold')
    plt.ylabel('Valores Reales', fontsize=14, weight='bold')
    plt.xticks(fontsize=12)
    plt.yticks(fontsize=12)
    plt.show()

def cargar_modelo_y_scaler():
    """Carga el modelo y el scaler si existen."""
    if os.path.exists(MODEL_PATH) and os.path.exists(SCALER_PATH):
        model = load_model(MODEL_PATH)
        scaler = joblib.load(SCALER_PATH)
        print("Modelo y scaler cargados correctamente.")
        return model, scaler
    return None, None

def ejecutar_cnn(df, n_splits=5):
    """
    Ejecuta la CNN con Stratified K-fold Cross-Validation o carga el modelo si existe.

    Args:
        df (pd.DataFrame): DataFrame con los datos.
        n_splits (int): Número de folds para la validación cruzada.

    Returns:
        tuple: modelo final, scaler
    """
    # Verificar si el modelo y el scaler ya están guardados
    model, scaler = cargar_modelo_y_scaler()
    if model is not None and scaler is not None:
        return model, scaler

    # Si no existen, entrenar el modelo
    establecer_semilla()
    start_time = time.time()

    # Preparar datos para validación cruzada
    bandas = df.iloc[:, 4:7]
    filtro = (bandas <= 0.35).all(axis=1)
    df_filtrado = df[filtro].copy()
    X = df_filtrado.iloc[:, 4:].values
    y = df_filtrado['Especie'].apply(lambda x: 0 if x == 'PIC' else 1).values
    print(f"Filas originales: {len(df)}, Filas después de filtrar: {len(df_filtrado)}")
    print(f"Forma de X: {X.shape}")

    # Inicializar StratifiedKFold
    skf = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=42)
    accuracies, losses, classification_reports = [], [], []

    fold = 1
    for train_index, test_index in skf.split(X, y):
        print(f"\nProcesando Fold {fold}/{n_splits}...")
        X_train, X_test = X[train_index], X[test_index]
        y_train, y_test = y[train_index], y[test_index]

        # Aplicar SMOTE
        smote = SMOTE(random_state=42)
        X_train_2d = X_train.reshape(X_train.shape[0], -1)
        X_train_smote, y_train_smote = smote.fit_resample(X_train_2d, y_train)
        X_train = X_train_smote.reshape(X_train_smote.shape[0], X_train.shape[1], 1)
        y_train = y_train_smote

        # Contar clases
        print(f'En y_train antes de SMOTE, 0: {Counter(y_train)[0]}, 1: {Counter(y_train)[1]}')
        print(f'En y_train después de SMOTE, 0: {Counter(y_train_smote)[0]}, 1: {Counter(y_train_smote)[1]}')
        print(f'En y_test, 0: {Counter(y_test)[0]}, 1: {Counter(y_test)[1]}')

        # Normalización (desactivada para pruebas)
        scaler = StandardScaler()
        # X_train = scaler.fit_transform(X_train.reshape(-1, X_train.shape[1])).reshape(X_train.shape)
        # X_test = scaler.transform(X_test.reshape(-1, X_test.shape[1])).reshape(X_test.shape[0], X_test.shape[1], 1)

        X_test = X_test.reshape(X_test.shape[0], X_test.shape[1], 1)
        print(f"Forma de X_train: {X_train.shape}, Forma de X_test: {X_test.shape}")

        # Crear y entrenar el modelo
        model = crear_modelo(input_shape=(X_train.shape[1], X_train.shape[2]))
        early_stopping = EarlyStopping(monitor='val_accuracy', patience=PATIENCE, restore_best_weights=True)
        history = model.fit(
            X_train, y_train, epochs=EPOCHS, batch_size=BATCH_SIZE,
            validation_data=(X_test, y_test), callbacks=[early_stopping], verbose=1
        )

        # Evaluar el modelo
        loss, accuracy = model.evaluate(X_test, y_test)
        y_pred_prob = model.predict(X_test)
        y_pred = (y_pred_prob > 0.5).astype("int32").flatten()
        accuracies.append(accuracy)
        losses.append(loss)
        classification_reports.append(classification_report(y_test, y_pred, output_dict=True))

        # Mostrar resultados
        mostrar_resultados(y_test, y_pred, loss, accuracy, f"CNN Fold {fold}", y_pred_prob, y_test)

        # Graficar métricas
        plt.figure(figsize=(12, 6))
        plt.subplot(1, 2, 1)
        plt.plot(history.history['accuracy'], label='Precisión Entrenamiento', color='darkblue', marker='o')
        plt.plot(history.history['val_accuracy'], label='Precisión Prueba', color='orange', marker='s')
        plt.title(f'Precisión Entrenamiento vs Prueba (Fold {fold})', fontsize=14)
        plt.xlabel('Época', fontsize=12)
        plt.ylabel('Precisión', fontsize=12)
        plt.legend(loc='upper left')

        plt.subplot(1, 2, 2)
        plt.plot(history.history['loss'], label='Pérdida Entrenamiento', color='crimson', linestyle='-.')
        plt.plot(history.history['val_loss'], label='Pérdida Prueba', color='darkgreen', linestyle='--')
        plt.title(f'Pérdida Entrenamiento vs Prueba (Fold {fold})', fontsize=14)
        plt.xlabel('Época', fontsize=12)
        plt.ylabel('Pérdida (Entropía Cruzada Binaria)', fontsize=12)
        plt.legend(loc='upper right')
        plt.tight_layout()
        plt.show()

        fold += 1

    # Mostrar métricas promedio
    print("\n=== Resultados de K-fold Cross-Validation ===")
    print(f"Precisión promedio: {np.mean(accuracies):.4f} ± {np.std(accuracies):.4f}")
    print(f"Pérdida promedio: {np.mean(losses):.4f} ± {np.std(losses):.4f}")
    avg_precision = np.mean([rep['weighted avg']['precision'] for rep in classification_reports])
    avg_recall = np.mean([rep['weighted avg']['recall'] for rep in classification_reports])
    avg_f1 = np.mean([rep['weighted avg']['f1-score'] for rep in classification_reports])
    print("\nInforme de clasificación promedio:")
    print(f"Precisión: {avg_precision:.4f}")
    print(f"Recall: {avg_recall:.4f}")
    print(f"F1-score: {avg_f1:.4f}")

    # Entrenar modelo final
    X_train, X_test, y_train, y_test, scaler = preparar_datos(df)
    model = crear_modelo(input_shape=(X_train.shape[1], X_train.shape[2]))
    early_stopping = EarlyStopping(monitor='val_accuracy', patience=PATIENCE, restore_best_weights=True)
    model.fit(
        X_train, y_train, epochs=EPOCHS, batch_size=BATCH_SIZE,
        validation_data=(X_test, y_test), callbacks=[early_stopping]
    )

    # Guardar el modelo y el scaler
    model.save(MODEL_PATH)
    joblib.dump(scaler, SCALER_PATH)
    print(f"Modelo guardado en: {MODEL_PATH}")
    print(f"Scaler guardado en: {SCALER_PATH}")

    end_time = time.time()
    print(f'Tiempo total de ejecución: {end_time - start_time:.2f} segundos')
    return model, scaler

import pandas as pd
import numpy as np
from collections import Counter
import openpyxl
from openpyxl.styles import Font, Alignment
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.metrics import classification_report, confusion_matrix

def comprobar_nuevos_datos(model, nuevos_datos, scaler, output_excel="predicciones_por_hoja.xlsx"):
    """
    Preprocesa nuevos datos, realiza predicciones y genera un Excel con la especie mayoritaria por hoja.

    Args:
        model: Modelo entrenado.
        nuevos_datos (pd.DataFrame): DataFrame con nuevos datos.
        scaler: Escalador usado en el entrenamiento.
        output_excel (str): Nombre del archivo Excel de salida.
    """
    print(f"Número de columnas en el DataFrame de nuevos datos: {nuevos_datos.shape[1]}")

    bandas_nuevas = nuevos_datos.iloc[:, 4:7]
    filtro_nuevas = (bandas_nuevas <= 0.35).all(axis=1)
    nuevos_datos_filtrados = nuevos_datos[filtro_nuevas].copy()
    print(f"Filas originales (nuevos datos): {len(nuevos_datos)}, Filas después de filtrar: {len(nuevos_datos_filtrados)}")

    if len(nuevos_datos_filtrados) == 0:
        print("Error: No hay datos después de filtrar. Verifica los datos de entrada.")
        return

    X_nuevos = nuevos_datos_filtrados.iloc[:, 4:].values

    if 'Especie' in nuevos_datos_filtrados.columns:
        y_nuevos = nuevos_datos_filtrados['Especie'].apply(lambda x: 0 if x == 'PIC' else 1).values
        hojas = nuevos_datos_filtrados['Hoja'].values
    else:
        print("Depuración: Columna 'Especie' no encontrada, omitiendo predicciones basadas en especie.")
        y_nuevos = np.zeros(len(nuevos_datos_filtrados))  # Valores por defecto
        hojas = nuevos_datos_filtrados.index.values  # Usar índices como placeholder

    print(f"Forma de X_nuevos antes de reshape: {X_nuevos.shape}")
    X_nuevos = X_nuevos.reshape((X_nuevos.shape[0], X_nuevos.shape[1]))
    X_nuevos = X_nuevos.reshape((X_nuevos.shape[0], X_nuevos.shape[1], 1)).astype(np.float32)
    print(f"Forma de X_nuevos después de reshape: {X_nuevos.shape}")
    print(f"Tipo de X_nuevos: {type(X_nuevos)}, dtype: {X_nuevos.dtype}")

    # Realizar predicciones
    y_pred_prob = model.predict(X_nuevos)
    y_pred = (y_pred_prob > 0.5).astype("int32").flatten()
    y_pred_especie = ['PIC' if pred == 0 else 'No PIC' for pred in y_pred]

    print("\nPredicciones para los nuevos datos (especies):")
    print(y_pred_especie)
    print("\nValores reales:")
    print(y_nuevos)
    print("\nInforme de clasificación para los nuevos datos:")
    if 'Especie' in nuevos_datos_filtrados.columns:
        print(classification_report(y_nuevos, y_pred))
    else:
        print("No se generó informe de clasificación porque no hay columna 'Especie'.")

    if 'Especie' in nuevos_datos_filtrados.columns:
        cm = confusion_matrix(y_nuevos, y_pred)
        plt.figure(figsize=(8, 6))
        sns.heatmap(cm, annot=True, fmt="d", cmap="YlGnBu", cbar=False,
                    annot_kws={"size": 16, "weight": "bold"}, linewidths=1.5, linecolor="black")
        plt.title('Matriz de Confusión - Nuevos Datos')
        plt.xlabel('Predicciones')
        plt.ylabel('Valores Reales')
        plt.xticks(fontsize=12)
        plt.yticks(fontsize=12)
        plt.show()

    predicciones_por_hoja = pd.DataFrame({
        'Hoja': hojas,
        'Especie_Predicha': y_pred_especie
    })

    especie_mayoritaria_por_hoja = {}
    for hoja in predicciones_por_hoja['Hoja'].unique():
        especies_hoja = predicciones_por_hoja[predicciones_por_hoja['Hoja'] == hoja]['Especie_Predicha']
        conteo = Counter(especies_hoja)
        especie_mayoritaria = conteo.most_common(1)[0][0]
        especie_mayoritaria_por_hoja[hoja] = especie_mayoritaria

    wb = openpyxl.Workbook()
    wb.remove(wb['Sheet'])

    ws_hojas = wb.create_sheet(title="Especies Mayoritarias por Hoja")
    ws_hojas['A1'] = 'Hoja'
    ws_hojas['B1'] = 'Especie Mayoritaria'
    ws_hojas['A1'].font = Font(bold=True)
    ws_hojas['B1'].font = Font(bold=True)
    ws_hojas['A1'].alignment = Alignment(horizontal='center')
    ws_hojas['B1'].alignment = Alignment(horizontal='center')

    row = 2
    for hoja, especie in especie_mayoritaria_por_hoja.items():
        ws_hojas[f'A{row}'] = hoja
        ws_hojas[f'B{row}'] = especie
        ws_hojas[f'A{row}'].alignment = Alignment(horizontal='center')
        ws_hojas[f'B{row}'].alignment = Alignment(horizontal='center')
        row += 1

    ws_predicciones = wb.create_sheet(title="Predicciones por Hoja")
    ws_predicciones['A1'] = 'Hoja'
    ws_predicciones['B1'] = 'Especie Predicha'
    ws_predicciones['A1'].font = Font(bold=True)
    ws_predicciones['B1'].font = Font(bold=True)
    ws_predicciones['A1'].alignment = Alignment(horizontal='center')
    ws_predicciones['B1'].alignment = Alignment(horizontal='center')

    row = 2
    for idx, (index, row_data) in enumerate(predicciones_por_hoja.iterrows(), start=row):
        ws_predicciones[f'A{idx}'] = row_data['Hoja']
        ws_predicciones[f'B{idx}'] = row_data['Especie_Predicha']
        ws_predicciones[f'A{idx}'].alignment = Alignment(horizontal='center')
        ws_predicciones[f'B{idx}'].alignment = Alignment(horizontal='center')

    # Ajustar ancho de columnas
    ws_hojas.column_dimensions['A'].width = 15
    ws_hojas.column_dimensions['B'].width = 20
    ws_predicciones.column_dimensions['A'].width = 15
    ws_predicciones.column_dimensions['B'].width = 20

    wb.save(output_excel)
    print(f"\nArchivo Excel generado: {output_excel}")

if __name__ == "__main__":
    data = pd.read_excel('../../Resultados/excel_2/TODOS/TODOS.xlsx')
    nuevos_datos = pd.read_excel('../../Resultados/PICUAL/Picual_4/3_Segmentacion_2/resultados_pic4_2.xlsx')
    comparar_con_pca(data.iloc[:, 4:].values, nuevos_datos.iloc[:, 4:].values)

    model, scaler = ejecutar_cnn(data)
    comprobar_nuevos_datos(model, nuevos_datos, scaler)